from openpyxl.reader.excel import load_workbook


class webData:
    """
    This class is used to extract the test data from TestData.xlsx sheet
    """

    def __init__(self):
        self.url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
        self.dashboardUrl = "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index"
        self.filename = "Data/TestData.xlsx"
        self.sheetName = "Sheet1"
        self.workbook = load_workbook(self.filename)
        self.sheet = self.workbook[self.sheetName]

    def rowcount(self):
        """
        This function return max rowcount till the data is present
        :return:
        """
        return self.sheet.max_row

    def readData(self, row, column):
        """
        This functions read the data in particular cell in excel
        :param row:
        :param column:
        :return:
        """
        return self.sheet.cell(row, column).value

    def writeData(self, row, column, data):
        """
        This functions write the data in particular cell in excel
        :param row:
        :param column:
        :param data:
        :return:
        """
        self.sheet.cell(row, column).value = data
        self.workbook.save(self.filename)

